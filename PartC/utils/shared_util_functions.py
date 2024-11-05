import requests
import time
import psutil
from openpyxl import Workbook, load_workbook

java_api_pid = 15640

def get_metrics(url, method_type, method_body):
    # Start time of api call
    start_time = time.time()

    # Run api call
    if(method_type == 'POST'):
        response = requests.post(url, json={"title": method_body["title"]})
    elif(method_type == "PUT"):
        response = requests.put(url, json={"title": method_body["title"]})
    elif(method_type == "DELETE"):
        response = requests.delete(url)

    # Retrieve the cpu usage and memory usage before the api call
    cpu_percent = psutil.cpu_percent(interval=0.1) 
    mem_percent = psutil.virtual_memory().percent

    # End time of api call
    end_time = time.time()   

    # Calculate the response time, cpu usage and memory usage
    response_time = end_time - start_time

    return response_time, cpu_percent, mem_percent


# Methods to create excel file and necessary shhets append data to excel file and close the workbook
def create_excel_file(file_path):
    workbook = Workbook()
    post_performance_data_sheet = workbook.active
    post_performance_data_sheet.title = "POST Performance Data"
    put_performance_data_sheet = workbook.create_sheet(title="PUT Performance Data")
    delete_performance_data_sheet = workbook.create_sheet(title="DELETE Performance Data")

    headers = (["Number of Objects", "Transaction Time 1 (s)", "Transaction Time 2 (s)", "Transaction Time 3 (s)", "Transaction Time 4 (s)", 
                "CPU Usage 1 (%)", "CPU Usage 2 (%)", "CPU Usage 3 (%)", "CPU Usage 4 (%)",
                "Memory Usage 1 (%)", "Memory Usage 2 (%)", "Memory Usage 3 (%)", "Memory Usage 4 (%)"])
    
    post_performance_data_sheet.append(headers)
    put_performance_data_sheet.append(headers)
    delete_performance_data_sheet.append(headers)
    
    workbook.save(file_path)

def append_data_to_excel(file_path, sheet_name, data):
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]    
    
    sheet.append(data)
    
    workbook.save(file_path)

def close_workbook(file_path):
    workbook = load_workbook(file_path)
    workbook.close()
    

